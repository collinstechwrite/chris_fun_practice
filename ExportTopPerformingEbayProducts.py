import sys
import urllib
import re
from datetime import datetime
import requests
import time
from random import randint, choice, sample, randrange, shuffle
from time import sleep
import os
from openpyxl import Workbook, load_workbook
from colorama import init, Fore, Back, Style
from io import BytesIO



init(convert=True)



errorcount = 0
countpromotions = 0
totalprice = 0.00
extractMostPopular = []
price = 0.00
favouritecount = 0











def increment():
    global countpromotions
    global totalprice
    global price
    
    if countpromotions >= 8:
        countpromotions = 0
    countpromotions = countpromotions+1
    totalprice = totalprice + price




print("CHECKING HOT EBAY USA TRENDS")
linkforpopular = ("https://www.wuanto.com/mostwatched/allkw/0/0") #This link has hot trends of things for sale in USA
zz = requests.get(linkforpopular)
populardata = zz.text
getlistofkeywords = re.findall('href="mostwatched/(.+?)/0/0', populardata) #extract a most sold keyword


for word in getlistofkeywords:        
    extractMostPopular.append(word) 


shuffle(extractMostPopular)


print(extractMostPopular)


myfilelist =[]

try:
    
    fh = open('myimportfile.txt')
    for line in fh:
        # in python 2
        # print line
        # in python 3
        myfilelist.append(line)
    fh.close()

    print("MY OWN FILE LIST")
    shuffle(myfilelist)
    print(myfilelist)


except FileNotFoundError:
    sys.exit("You need to create a file called myimportfile.txt and place it in the folder with .exe")



















#WORKING WITH DATETIME TO TIME STAMP SPREADSHEET LATER

now = datetime.now() # current date and time
keyword = ""
dateforspreadsheet = now.strftime("%m/%d/%Y")
yearforspreadsheet = now.strftime("%Y")
monthforspreadsheet = now.strftime("%B")
timeforspreadsheet = now.strftime("%H:%M:%S")




workbook = Workbook()
sheet = workbook.active

Row_Incrementer = 1

sheet["A" + str(Row_Incrementer)] = "itemID"
sheet["B" + str(Row_Incrementer)] = "title"
sheet["C" + str(Row_Incrementer)] = "price"
sheet["D" + str(Row_Incrementer)] = "watchers"
sheet["E" + str(Row_Incrementer)] = "mostSold"
sheet["F" + str(Row_Incrementer)] = "mostSoldinhours"
sheet["G" + str(Row_Incrementer)] = "keyword"
sheet["H" + str(Row_Incrementer)] = "date"
sheet["I" + str(Row_Incrementer)] = "year"
sheet["J" + str(Row_Incrementer)] = "month"
sheet["K" + str(Row_Incrementer)] = "time"



workbook.save(filename="analysis.xlsx")



your_html_file_name = input("e.g. WANT TO CREATE A HTML FILE? WRITE A FILENAME TO SAVE LIKE myhtmlfile.txt")
new_string = '<html><head><title>Check out ebay</title></head><body><h1>eBay Items For Sale</h1><p>Check out these ebay items for sale</p>'
opened_file = open(your_html_file_name, 'a')
opened_file.write(new_string)
opened_file.close()





















def pricefix(qb):

    try:
        pricef = float(qb)
        return pricef

    except Exception:

        qb = 0
        return qb
        
        



def removeap(z):
        z = z.replace(",", "")
        return z


def AnalyseAnItemIDForTitle(analyse):
    
    try:
        linkforfindinghowmanysold = ("https://www.ebay.com/itm/" + analyse)
        h = requests.get(linkforfindinghowmanysold)
        mostsolddata = h.text
        extractMostSoldonEbay = re.findall('<title>(.+?)  | eBay</title>', mostsolddata)
        apple = str(extractMostSoldonEbay)
        az = apple[2:-2]

        if az == "":
            linkforfindinghowmanysold = ("https://www.ebay.com/itm/" + analyse)
            h = requests.get(linkforfindinghowmanysold)
            mostsolddata = h.text
            extractMostSoldonEbay = re.findall('<meta name="description" content="(.+?)" />', mostsolddata)
            apple = str(extractMostSoldonEbay)
            az = apple[2:-2]
            return az
        

        return az
    except Exception:
        abz = 0
        return abz




def AnalyseAnItemID(analyse):


    try:
        linkforfindinghowmanysold = ("https://www.ebay.com/itm/" + analyse)
        h = requests.get(linkforfindinghowmanysold)
        mostsolddata = h.text
        extractMostSoldonEbay = re.findall('">(.+?) sold</a></span>', mostsolddata)
        apple = str(extractMostSoldonEbay)
        apple = apple.replace(',', '')
        ap = apple[2:-2]
        az = int(ap)

        return az
    except Exception:
        abz = 0
        return abz


def AnalyseAnItemIDForhours(analyse2):
    try:
        linkforfindinghowmanysold = ("https://www.ebay.com/itm/" + analyse2)
        h = requests.get(linkforfindinghowmanysold)
        mostsolddata = h.text
        extractMostSoldonEbay = re.findall('>(.+?) sold in last 24 hours</span>', mostsolddata)
        apple = str(extractMostSoldonEbay)
        apple = apple.replace(',', '')
        ap = apple[2:-2]
        az = int(ap)
  
        return az
    except Exception:
        abz = 0
        return abz



def AnalyseItemIDForPrice(analyse):


    try:
        linkforfindinghowmanysold = ("https://www.ebay.com/itm/" + analyse)
        h = requests.get(linkforfindinghowmanysold)
        mostsolddata = h.text
        extractMostSoldonEbay = re.findall('">US(.+?)</span>', mostsolddata)
        apple = str(extractMostSoldonEbay)
        apple = apple.replace('$', '')
        apple = apple.replace(' ', '')
        apple = apple.replace('/ea','')
        ap = apple[2:-2]
        az = ap
       
        return az
    except Exception:
        abz = 0
        return abz




def AnalyseAnItemIDForWatchers(analyse2):
    try:
        linkforfindinghowmanysold = ("https://www.ebay.com/itm/" + analyse2)
        h = requests.get(linkforfindinghowmanysold)
        mostsolddata = h.text
        extractMostSoldonEbay = re.findall('defaultWatchCount":(.+?),', mostsolddata)
        apple = str(extractMostSoldonEbay)
        apple = apple.replace(',', '')
        ap = apple[2:-2]
        az = int(ap)
   
        return az
    except Exception:
        abz = 0
        return abz






def unescape(s):
    s = s.replace("&lt;", "<")
    s = s.replace("&gt;", ">")
    # this has to be last:
    s = s.replace("&amp;", "&")
    s = s.replace("Collection of products named ", "")
    s = s.replace("Advanced Search" , "")
    s = s.replace("Verify site&#39;s SSL certificate", "")
    s = s.replace("&#39;", "")
    return s

def unescape2(s):
    s = s.replace(" ", "+")
    return s

def unescape4(s):
    s = s.replace(" ", "+")
    s = s.replace("\n", "")
    return s


def unescape5(s):
    s = s.replace(",", "")
    return s


def unescape6(s):
    s = s.replace("&apos;", "'")
    s = s.replace("&amp;", "&")
    s = s.replace('&quot;', '"')
    s = s.replace("', '",",")
    return s

def row_increment(itemID,title,price,watchers,mostSold,mostSoldinhours,pear,dateforspreadsheet,monthforspreadsheet,timeforspreadsheet,myebayurl):

    global Row_Incrementer
    Row_Incrementer += 1


    workbook = load_workbook(filename="analysis.xlsx")
    sheet = workbook.active

    sheet["A" + str(Row_Incrementer)] = itemID
    sheet["B" + str(Row_Incrementer)] = title
    sheet["C" + str(Row_Incrementer)] = price
    sheet["D" + str(Row_Incrementer)] = watchers
    sheet["E" + str(Row_Incrementer)] = mostSold
    sheet["F" + str(Row_Incrementer)] = mostSoldinhours


    sheet["G" + str(Row_Incrementer)] = pear
    sheet["H" + str(Row_Incrementer)] = dateforspreadsheet
    sheet["I" + str(Row_Incrementer)] = yearforspreadsheet
    sheet["J" + str(Row_Incrementer)] = monthforspreadsheet
    sheet["K" + str(Row_Incrementer)] = timeforspreadsheet
    sheet["L" + str(Row_Incrementer)] = myebayurl

    workbook.save(filename="analysis.xlsx")




def html_file_update(your_html_file_name,title,yourebaycampaignID,price,itemID):

    myebayurl = "http://rover.ebay.com/rover/1/711-53200-19255-0/1?icep_ff3=2&pub=5575036718&toolid=10001&campid=" + yourebaycampaignID +"&customid=&icep_item=" + itemID + "&ipn=psmain&icep_vectorid=229466&kwid=902099&mtid=824&kw=lg"

    #piece of code working with title
    new_string = '<H2>' + title + '</H2>'
    opened_file = open(your_html_file_name, 'a')
    opened_file.write(new_string)
    opened_file.close()


    #piece of code working with myebayurl
    new_string = "<table border='0' cellpadding='8'> <tr><td> <a href=" + myebayurl + "' target='_blank'>"
    opened_file = open(your_html_file_name, 'a')
    opened_file.write(new_string)
    opened_file.close()

    
    #write new piece of code to extract thumbnail picture

    linkforfindinghowmanysold = ("https://www.ebay.com/itm/" + itemID)
    h = requests.get(linkforfindinghowmanysold)
    mostsolddata = h.text
    extractMostSoldonEbay = re.findall('<meta name="twitter:image" content="https://i.ebayimg.com/images/g/(.+?).jpg" />', mostsolddata)
    extractMostSoldonEbay = extractMostSoldonEbay[0]
  


    #write new piece of code to extract thumbnail picture
    new_string = "<img src='https://i.ebayimg.com/thumbs/images/g/" + extractMostSoldonEbay + ".jpg' border='0'/></a></td><td><strong><H3>$" + price + "</H3></strong>"
    opened_file = open(your_html_file_name, 'a')
    opened_file.write(new_string)
    opened_file.close()


     
    #piece of code working with price
    new_string = "<br><H2>Buy It Now for only: $" + price + "</H2>"
    opened_file = open(your_html_file_name, 'a')
    opened_file.write(new_string)
    opened_file.close()



    #piece of code working with myebayurl
    new_string = "<br><a href='" + myebayurl + "' target='_blank'><div style="
    
    opened_file = open(your_html_file_name, 'a')
    opened_file.write(new_string)
    opened_file.close()



    #piece of code completing button
    new_string = 'display: inline-block; border: 3px solid #f41f1f; border-radius: 3px; padding: 8px; background: linear-gradient(to bottom,#f2e6e6,#f5afaf);">Buy Now</div></a> </table><br></br>'
    
    opened_file = open(your_html_file_name, 'a')
    opened_file.write(new_string)
    opened_file.close()



#CODE FOR REMOVING UTF FORMAT
def BMP(s):
    return "".join((i if ord(i) < 10000 else '\ufffd' for i in s))


def ownfile():
    
    try:


        global myfilelist
        
        myset_data = []
        
        for fruit in myfilelist:


            pear = unescape2(fruit)
            print(pear)
            print('{:8} {:8} {:5} {:10} {:14} {:10}'.format('Watch', 'Sold', 'Hour', 'Price','ItemID','Title'))
            linkforprocessing = ("http://rest.ebay.com/epn/v1/find/item.rss?keyword="+ pear +"&sortOrder=BestMatch&programid=1&campaignid=5337424366&toolid=10039&minPrice="+ priceinput +".0&listingType1=All&topRatedSeller=true&feedType=rss&lgeo=1")

            f = requests.get(linkforprocessing)
             
            
            websiteData = f.text
            
            

            extractEbayItemIDFromWebsite = re.findall ('<guid>(.+?)</guid>', websiteData)




            global price
            for itemID in extractEbayItemIDFromWebsite:
            
                title = AnalyseAnItemIDForTitle(itemID)
                price = AnalyseItemIDForPrice(itemID)
                title = BMP(title)
                title = unescape6(title)
                title = title[0:81]
                price = unescape5(price)
                price = pricefix(price)
                price = float(price)
                mostSold = AnalyseAnItemID(itemID)
                mostSoldinhours = AnalyseAnItemIDForhours(itemID)
                watchers = AnalyseAnItemIDForWatchers(itemID)
                

                print('{:*^8} {:_>8} {:*^5} {:_>10.2f} {:_^14} {:_<10}'.format(watchers, mostSold, mostSoldinhours, price, itemID, title))


                if price >= float(priceinput) and watchers >= int(watchersinput) and mostSold >= int(mostSoldinput) and mostSoldinhours >= int(mostSoldinhoursvar):

                    
                    
                    global countpromotions
                    global totalprice
                    increment()
                    

                    if mostSold > 2000 and mostSoldinhours > 2:

                        if itemID in myset_data:
                            print("Skipping Duplicate")
                            break
                        myset_data.append(itemID)

                        title = str(title)
                        price = '{0:.2f}'.format(price)
                        price = str(price)
                        mostSold = str(mostSold)
                        mostSoldinhours = str(mostSoldinhours)
                        watchers = str(watchers)
                        itemID = str(itemID)

                        now = datetime.now() # current date and time
                        keyword = pear
                        dateforspreadsheet = now.strftime("%m/%d/%Y")
                        yearforspreadsheet = now.strftime("%Y")
                        monthforspreadsheet = now.strftime("%B")
                        timeforspreadsheet = now.strftime("%H:%M:%S")

                        
                        print(str(title), file=open(myfile, 'a'))
                        fileforappending.close()
                        print("appending ITEM ID: " + itemID + " : " + title + " data to file")
                        html_file_update(your_html_file_name,title,yourebaycampaignID,price,itemID)



                        myebayurl = "http://rover.ebay.com/rover/1/711-53200-19255-0/1?icep_ff3=2&pub=5575036718&toolid=10001&campid=" + yourebaycampaignID +"&customid=&icep_item=" + itemID + "&ipn=psmain&icep_vectorid=229466&kwid=902099&mtid=824&kw=lg"

                        row_increment(itemID,title,price,watchers,mostSold,mostSoldinhours,pear,dateforspreadsheet,monthforspreadsheet,timeforspreadsheet,myebayurl)

                    
                        

                        #CHECK EBAY ITEM CONVERSION URL AND MOSTSOLDINHOUR TEXT

                        
                        mystr = title
                        wordList = re.sub("[^\w]", " ",  mystr).split()

                        NEWWORDLIST = (sample(wordList, 2))

                        
                        
                        
                        tag1 = NEWWORDLIST[0]
                        tag2 = NEWWORDLIST[1]

                        

                        MyRandomPhrases = ("Thousands sold","Hot Product","Trending","Selling Fast","Fantastic offer","Why wait?","Trusted seller!","Top seller!","Want","Try","Terrific","Superb","Super","Nice","Looking for","Great Deal","Give","Gift Item","Best Ever","Best","Affordable","A fabulous gift")
                        MyRandomPrices = ("Yours for","Only","Just","Buy it now!","On sale","For","Get it now","Buy now","Going for","Best price","Bargain","Selling for")
                        num = randrange(0,20)
                        num2 = randrange(0,10)
                    
                        print("{} #{} #{} #{}, Sold over {}, {} sold in last hour. {} ${} http://rover.ebay.com/rover/1/711-53200-19255-0/1?icep_ff3=2&pub=5575036718&toolid=10001&campid={}&customid=&icep_item={}&ipn=psmain&icep_vectorid=229466&kwid=902099&mtid=824&kw=lg ".format(MyRandomPhrases[num],title,str(tag1),str(tag2),mostSold,mostSoldinhours,MyRandomPrices[num2],price,yourebaycampaignID,itemID))
                        
                        print('NEXT PROCESS')
                        


                        
                        print(errorcount)
                        print(countpromotions)
                        print(totalprice)

                        if errorcount >= 3:
                            sys.exit("TOO MANY ERRORS HAVE OCCURED. EXITING TWEETER. HAVE A NICE DAY! CHECK YOUR TWITTER ISN'T LOCKED. TRY AGAIN IN A FEW HOURS.")
                            
                        if countpromotions >= 7:

                            break
     


                    elif mostSold > 0 and mostSoldinhours > 0:

                        if itemID in myset_data:
                            print("Skipping Duplicate")
                            break
                        myset_data.append(itemID)

                        title = str(title)
                        price = '{0:.2f}'.format(price)
                        price = str(price)
                        mostSold = str(mostSold)
                        mostSoldinhours = str(mostSoldinhours)
                        watchers = str(watchers)
                        itemID = str(itemID)

                        now = datetime.now() # current date and time
                        keyword = pear
                        dateforspreadsheet = now.strftime("%m/%d/%Y")
                        yearforspreadsheet = now.strftime("%Y")
                        monthforspreadsheet = now.strftime("%B")
                        timeforspreadsheet = now.strftime("%H:%M:%S")

                        
                        print(str(title), file=open(myfile, 'a'))
                        fileforappending.close()
                        print("appending ITEM ID: " + itemID + " : " + title + " data to file")
                        html_file_update(your_html_file_name,title,yourebaycampaignID,price,itemID)


                        myebayurl = "http://rover.ebay.com/rover/1/711-53200-19255-0/1?icep_ff3=2&pub=5575036718&toolid=10001&campid=" + yourebaycampaignID +"&customid=&icep_item=" + itemID + "&ipn=psmain&icep_vectorid=229466&kwid=902099&mtid=824&kw=lg"



                        row_increment(itemID,title,price,watchers,mostSold,mostSoldinhours,pear,dateforspreadsheet,monthforspreadsheet,timeforspreadsheet,myebayurl)

                    

                        #CHECK EBAY ITEM CONVERSION URL AND MOSTSOLDINHOUR TEXT

                        
                        mystr = title
                        wordList = re.sub("[^\w]", " ",  mystr).split()

                        NEWWORDLIST = (sample(wordList, 2))

                        
                        
                        
                        tag1 = NEWWORDLIST[0]
                        tag2 = NEWWORDLIST[1]

                        

                        MyRandomPhrases = ("Want","Try","Terrific","Superb","Super","Nice","Looking for","Great Deal","Give","Gift Item","Best Ever","Best","Affordable","A fabulous gift")
                        MyRandomPrices = ("Yours for","Only","Just","Buy it now!","On sale","For","Get it now","Buy now","Going for","Best price","Bargain","Selling for")
                        num = randrange(0,12)
                        num2 = randrange(0,10)
                    
                        print("{} #{} #{} #{}, Sold over {}, {} sold in last hour. {} ${} http://rover.ebay.com/rover/1/711-53200-19255-0/1?icep_ff3=2&pub=5575036718&toolid=10001&campid={}&customid=&icep_item={}&ipn=psmain&icep_vectorid=229466&kwid=902099&mtid=824&kw=lg ".format(MyRandomPhrases[num],title,str(tag1),str(tag2),mostSold,mostSoldinhours,MyRandomPrices[num2],price,yourebaycampaignID,itemID))
                        
                        print('NEXT PROCESS')
                        print(errorcount)
                        print(countpromotions)
                        print(totalprice)

                        if errorcount >= 3:
                            sys.exit("TOO MANY ERRORS HAVE OCCURED. EXITING TWEETER. HAVE A NICE DAY! CHECK YOUR TWITTER ISN'T LOCKED. TRY AGAIN IN A FEW HOURS.")
                            
                        if countpromotions >= 7:

                            break
                        



                    elif mostSold == 0:

                        if "apple" or "Apple" in title:


                            if itemID in myset_data:
                                print("Skipping Duplicate")
                                break
                            myset_data.append(itemID)

                
                            title = str(title)
                            price = '{0:.2f}'.format(price)
                            price = str(price)
                            mostSold = str(mostSold)
                            mostSoldinhours = str(mostSoldinhours)
                            watchers = str(watchers)
                            itemID = str(itemID)

                            now = datetime.now() # current date and time
                            keyword = pear
                            dateforspreadsheet = now.strftime("%m/%d/%Y")
                            yearforspreadsheet = now.strftime("%Y")
                            monthforspreadsheet = now.strftime("%B")
                            timeforspreadsheet = now.strftime("%H:%M:%S")

                            
                            print(str(title), file=open(myfile, 'a'))
                            fileforappending.close()
                            print("appending ITEM ID: " + itemID + " : " + title + " data to file")
                            html_file_update(your_html_file_name,title,yourebaycampaignID,price,itemID)

                            
                            myebayurl = "http://rover.ebay.com/rover/1/711-53200-19255-0/1?icep_ff3=2&pub=5575036718&toolid=10001&campid=" + yourebaycampaignID +"&customid=&icep_item=" + itemID + "&ipn=psmain&icep_vectorid=229466&kwid=902099&mtid=824&kw=lg"


                            row_increment(itemID,title,price,watchers,mostSold,mostSoldinhours,pear,dateforspreadsheet,monthforspreadsheet,timeforspreadsheet,myebayurl)


                            
                            mystr = title
                            wordList = re.sub("[^\w]", " ",  mystr).split()

                            NEWWORDLIST = (sample(wordList, 2))

                            
                            
                            
                            tag1 = NEWWORDLIST[0]
                            tag2 = NEWWORDLIST[1]
                            #RANDOM TECH PHRASES
                            MyRandomPhrases = ("Want","Try","Terrific","Superb","Super","Nice","Looking for","Great Deal","Give","Gift Item","Best Ever","Best","Affordable","A fabulous gift")
                            MyRandomPrices = ("Yours for","Only","Just","Buy it now!","On sale","For","Get it now","Buy now","Going for","Best price","Bargain","Selling for")
                            num = randrange(0,12)
                            num2 = randrange(0,10)
                        
                            print("{} #{} #{} #{}, {} Watchers, {} ${} http://rover.ebay.com/rover/1/711-53200-19255-0/1?icep_ff3=2&pub=5575036718&toolid=10001&campid={}&customid=&icep_item={}&ipn=psmain&icep_vectorid=229466&kwid=902099&mtid=824&kw=lg ".format(MyRandomPhrases[num],title,str(tag1),str(tag2),watchers,MyRandomPrices[num2],price,yourebaycampaignID,itemID))

                            
                            print('NEXT PROCESS')
                            print(errorcount)
                            print(countpromotions)
                            print(totalprice)


                            
                            
                            if errorcount >= 3:
                                sys.exit("TOO MANY ERRORS HAVE OCCURED. EXITING TWEETER. HAVE A NICE DAY! CHECK YOUR TWITTER ISN'T LOCKED. TRY AGAIN IN A FEW HOURS.")

                            if countpromotions >= 7:
                                
                                break
                    

                        elif "antique" or "Antique" or "antique," or "Antique," in title:


                            if itemID in myset_data:
                                print("Skipping Duplicate")
                                break
                            myset_data.append(itemID)



                            title = str(title)
                            price = '{0:.2f}'.format(price)
                            price = str(price)
                            mostSold = str(mostSold)
                            mostSoldinhours = str(mostSoldinhours)
                            watchers = str(watchers)
                            itemID = str(itemID)

                            now = datetime.now() # current date and time
                            keyword = pear
                            dateforspreadsheet = now.strftime("%m/%d/%Y")
                            yearforspreadsheet = now.strftime("%Y")
                            monthforspreadsheet = now.strftime("%B")
                            timeforspreadsheet = now.strftime("%H:%M:%S")


                            
                            print(str(title), file=open(myfile, 'a'))
                            fileforappending.close()
                            print("appending ITEM ID: " + itemID + " : " + title + " data to file")
                            html_file_update(your_html_file_name,title,yourebaycampaignID,price,itemID)

                            myebayurl = "http://rover.ebay.com/rover/1/711-53200-19255-0/1?icep_ff3=2&pub=5575036718&toolid=10001&campid=" + yourebaycampaignID +"&customid=&icep_item=" + itemID + "&ipn=psmain&icep_vectorid=229466&kwid=902099&mtid=824&kw=lg"


                            row_increment(itemID,title,price,watchers,mostSold,mostSoldinhours,pear,dateforspreadsheet,monthforspreadsheet,timeforspreadsheet,myebayurl)
                            
                            
                            mystr = title
                            wordList = re.sub("[^\w]", " ",  mystr).split()

                            NEWWORDLIST = (sample(wordList, 2))

                            
                            
                            
                            tag1 = NEWWORDLIST[0]
                            tag2 = NEWWORDLIST[1]


                            #RANDOM ANTIQUE
                            MyRandomPhrases = ("Rare","Collectable","Fine","Great find","Collector peice","Seeking","Looking for","Great Deal","Give","Gift Item","Best Ever","Best","Affordable","A fabulous gift")
                            MyRandomPrices = ("Yours for","Only","Just","Buy it now!","On sale","For","Get it now","Buy now","Going for","Best price","Bargain","Selling for")
                            num = randrange(0,12)
                            num2 = randrange(0,10)
                        
                            print("{} #{} #{} #{}, {} Watchers, {} ${} http://rover.ebay.com/rover/1/711-53200-19255-0/1?icep_ff3=2&pub=5575036718&toolid=10001&campid={}&customid=&icep_item={}&ipn=psmain&icep_vectorid=229466&kwid=902099&mtid=824&kw=lg ".format(MyRandomPhrases[num],title,str(tag1),str(tag2),watchers,MyRandomPrices[num2],price,yourebaycampaignID,itemID))

                            
                            print('NEXT PROCESS')
                            print(errorcount)
                            print(countpromotions)
                            print(totalprice)
  
                            if errorcount >= 3:
                                sys.exit("TOO MANY ERRORS HAVE OCCURED. EXITING TWEETER. HAVE A NICE DAY! CHECK YOUR TWITTER ISN'T LOCKED. TRY AGAIN IN A FEW HOURS.")

                            if countpromotions >= 7:
                                
                                break

                        
                        else:


                            if itemID in myset_data:
                                print("Skipping Duplicate")
                                break
                            myset_data.append(itemID)



                            title = str(title)
                            price = '{0:.2f}'.format(price)
                            price = str(price)
                            mostSold = str(mostSold)
                            mostSoldinhours = str(mostSoldinhours)
                            watchers = str(watchers)
                            itemID = str(itemID)

                            now = datetime.now() # current date and time
                            keyword = pear
                            dateforspreadsheet = now.strftime("%m/%d/%Y")
                            yearforspreadsheet = now.strftime("%Y")
                            monthforspreadsheet = now.strftime("%B")
                            timeforspreadsheet = now.strftime("%H:%M:%S")



                            
                            print(str(title), file=open(myfile, 'a'))
                            fileforappending.close()
                            print("appending ITEM ID: " + itemID + " : " + title + " data to file")
                            html_file_update(your_html_file_name,title,yourebaycampaignID,price,itemID)


                            workbook = load_workbook(filename="analysis.xlsx")
                            sheet = workbook.active


                            myebayurl = "http://rover.ebay.com/rover/1/711-53200-19255-0/1?icep_ff3=2&pub=5575036718&toolid=10001&campid=" + yourebaycampaignID +"&customid=&icep_item=" + itemID + "&ipn=psmain&icep_vectorid=229466&kwid=902099&mtid=824&kw=lg"


                            row_increment(itemID,title,price,watchers,mostSold,mostSoldinhours,pear,dateforspreadsheet,monthforspreadsheet,timeforspreadsheet,myebayurl)
                        
                            workbook.save(filename="analysis.xlsx")
                            
                            mystr = title
                            wordList = re.sub("[^\w]", " ",  mystr).split()

                            NEWWORDLIST = (sample(wordList, 2))

                            
                            
                            
                            tag1 = NEWWORDLIST[0]
                            tag2 = NEWWORDLIST[1]



                            MyRandomPhrases = ("Want","Try","Terrific","Superb","Super","Nice","Looking for","Great Deal","Give","Gift Item","Best Ever","Best","Affordable","A fabulous gift")
                            MyRandomPrices = ("Yours for","Only","Just","Buy it now!","On sale","For","Get it now","Buy now","Going for","Best price","Bargain","Selling for")
                            num = randrange(0,12)
                            num2 = randrange(0,10)
                        
                            print("{} #{} #{} #{}, {} Watchers, {} ${} http://rover.ebay.com/rover/1/711-53200-19255-0/1?icep_ff3=2&pub=5575036718&toolid=10001&campid={}&customid=&icep_item={}&ipn=psmain&icep_vectorid=229466&kwid=902099&mtid=824&kw=lg ".format(MyRandomPhrases[num],title,str(tag1),str(tag2),watchers,MyRandomPrices[num2],price,yourebaycampaignID,itemID))

                            
                            print('NEXT PROCESS')
                            print(errorcount)
                            print(countpromotions)
                            print(totalprice)

                            if errorcount >= 3:
                                sys.exit("TOO MANY ERRORS HAVE OCCURED. EXITING TWEETER. HAVE A NICE DAY! CHECK YOUR TWITTER ISN'T LOCKED. TRY AGAIN IN A FEW HOURS.")

                            if countpromotions >= 7:
                                
                                break

                    elif mostSold >= 0 and mostSoldinhours  == 0:


                        if itemID in myset_data:
                            print("Skipping Duplicate")
                            break
                        myset_data.append(itemID)

                        
                        title = str(title)
                        price = '{0:.2f}'.format(price)
                        price = str(price)
                        mostSold = str(mostSold)
                        mostSoldinhours = str(mostSoldinhours)
                        watchers = str(watchers)
                        itemID = str(itemID)

                        now = datetime.now() # current date and time
                        keyword = pear
                        dateforspreadsheet = now.strftime("%m/%d/%Y")
                        yearforspreadsheet = now.strftime("%Y")
                        monthforspreadsheet = now.strftime("%B")
                        timeforspreadsheet = now.strftime("%H:%M:%S")


                        
                        print(str(title), file=open(myfile, 'a'))
                        fileforappending.close()
                        print("appending ITEM ID: " + itemID + " : " + title + " data to file")
                        html_file_update(your_html_file_name,title,yourebaycampaignID,price,itemID)

                        myebayurl = "http://rover.ebay.com/rover/1/711-53200-19255-0/1?icep_ff3=2&pub=5575036718&toolid=10001&campid=" + yourebaycampaignID +"&customid=&icep_item=" + itemID + "&ipn=psmain&icep_vectorid=229466&kwid=902099&mtid=824&kw=lg"


                        row_increment(itemID,title,price,watchers,mostSold,mostSoldinhours,pear,dateforspreadsheet,monthforspreadsheet,timeforspreadsheet,myebayurl)

                    

                        #CHECK EBAY ITEM CONVERSION URL AND MOSTSOLDINHOUR TEXT

                        
                        mystr = title
                        wordList = re.sub("[^\w]", " ",  mystr).split()

                        NEWWORDLIST = (sample(wordList, 2))

                        
                        
                        
                        tag1 = NEWWORDLIST[0]
                        tag2 = NEWWORDLIST[1]

                        MyRandomPhrases = (myhashtags + "Want","Try","Terrific","Superb","Super","Nice","Looking for","Great Deal","Give","Gift Item","Best Ever","Best","Affordable","A fabulous gift")
                        num = randrange(0,13)
                    
                        print("{} #{} #{} #{}, {} Watchers, http://rover.ebay.com/rover/1/711-53200-19255-0/1?icep_ff3=2&pub=5575036718&toolid=10001&campid={}&customid=&icep_item={}&ipn=psmain&icep_vectorid=229466&kwid=902099&mtid=824&kw=lg ".format(MyRandomPhrases[num],title,str(tag1),str(tag2),watchers,yourebaycampaignID,itemID))
                       
                        print('NEXT PROCESS')

                        print(errorcount)
                        print(countpromotions)
                        print(totalprice)
                        
                        if errorcount >= 3:
                            sys.exit("TOO MANY ERRORS HAVE OCCURED. EXITING TWEETER. HAVE A NICE DAY! CHECK YOUR TWITTER ISN'T LOCKED. TRY AGAIN IN A FEW HOURS.")
                        

                        if countpromotions >= 7:

                            break
                                            
                    
                    elif mostSold == 0 and mostSoldinhours  == 0:


                        if itemID in myset_data:
                            print("Skipping Duplicate")
                            break
                        myset_data.append(itemID)
                        
                        title = str(title)
                        price = '{0:.2f}'.format(price)
                        price = str(price)
                        mostSold = str(mostSold)
                        mostSoldinhours = str(mostSoldinhours)
                        watchers = str(watchers)
                        itemID = str(itemID)

                        now = datetime.now() # current date and time
                        keyword = pear
                        dateforspreadsheet = now.strftime("%m/%d/%Y")
                        yearforspreadsheet = now.strftime("%Y")
                        monthforspreadsheet = now.strftime("%B")
                        timeforspreadsheet = now.strftime("%H:%M:%S")


                        
                        print(str(title), file=open(myfile, 'a'))
                        fileforappending.close()
                        print("appending ITEM ID: " + itemID + " : " + title + " data to file")
                        html_file_update(your_html_file_name,title,yourebaycampaignID,price,itemID)

                        myebayurl = "http://rover.ebay.com/rover/1/711-53200-19255-0/1?icep_ff3=2&pub=5575036718&toolid=10001&campid=" + yourebaycampaignID +"&customid=&icep_item=" + itemID + "&ipn=psmain&icep_vectorid=229466&kwid=902099&mtid=824&kw=lg"

                        row_increment(itemID,title,price,watchers,mostSold,mostSoldinhours,pear,dateforspreadsheet,monthforspreadsheet,timeforspreadsheet,myebayurl)

                        #CHECK EBAY ITEM CONVERSION URL AND MOSTSOLDINHOUR TEXT

                        
                        mystr = title
                        wordList = re.sub("[^\w]", " ",  mystr).split()

                        NEWWORDLIST = (sample(wordList, 2))

                        
                        
                        
                        tag1 = NEWWORDLIST[0]
                        tag2 = NEWWORDLIST[1]

                        MyRandomPhrases = (myhashtags + "Want","Try","Terrific","Superb","Super","Nice","Looking for","Great Deal","Give","Gift Item","Best Ever","Best","Affordable","A fabulous gift")
                        num = randrange(0,13)
                    
                        print("{} #{} #{} #{}, {} Watchers, http://rover.ebay.com/rover/1/711-53200-19255-0/1?icep_ff3=2&pub=5575036718&toolid=10001&campid={}&customid=&icep_item={}&ipn=psmain&icep_vectorid=229466&kwid=902099&mtid=824&kw=lg ".format(MyRandomPhrases[num],title,str(tag1),str(tag2),watchers,yourebaycampaignID,itemID))
                        
                        print('NEXT PROCESS')

                        print(errorcount)
                        print(countpromotions)
                        print(totalprice)
                        
                        if errorcount >= 3:
                            sys.exit("TOO MANY ERRORS HAVE OCCURED. EXITING TWEETER. HAVE A NICE DAY! CHECK YOUR TWITTER ISN'T LOCKED. TRY AGAIN IN A FEW HOURS.")
                        

                        if countpromotions >= 7:

                            break
                        

        
           
                    else:

                        if itemID in myset_data:
                            print("Skipping Duplicate")
                            break
                        myset_data.append(itemID)
            
                        title = str(title)
                        price = '{0:.2f}'.format(price)
                        price = str(price)
                        mostSold = str(mostSold)
                        mostSoldinhours = str(mostSoldinhours)
                        watchers = str(watchers)
                        itemID = str(itemID)

                        now = datetime.now() # current date and time
                        keyword = pear
                        dateforspreadsheet = now.strftime("%m/%d/%Y")
                        yearforspreadsheet = now.strftime("%Y")
                        monthforspreadsheet = now.strftime("%B")
                        timeforspreadsheet = now.strftime("%H:%M:%S")

                        
                        print(str(title), file=open(myfile, 'a'))
                        fileforappending.close()
                        print("appending ITEM ID: " + itemID + " : " + title + " data to file")
                        html_file_update(your_html_file_name,title,yourebaycampaignID,price,itemID)

                        myebayurl = "http://rover.ebay.com/rover/1/711-53200-19255-0/1?icep_ff3=2&pub=5575036718&toolid=10001&campid=" + yourebaycampaignID +"&customid=&icep_item=" + itemID + "&ipn=psmain&icep_vectorid=229466&kwid=902099&mtid=824&kw=lg"


                        row_increment(itemID,title,price,watchers,mostSold,mostSoldinhours,pear,dateforspreadsheet,monthforspreadsheet,timeforspreadsheet,myebayurl)

                        #CHECK EBAY ITEM CONVERSION URL AND MOSTSOLDINHOUR TEXT

                        mystr = title
                        wordList = re.sub("[^\w]", " ",  mystr).split()

                        NEWWORDLIST = (sample(wordList, 2))

                        
                        
                        
                        tag1 = NEWWORDLIST[0]
                        tag2 = NEWWORDLIST[1]
                        



                        MyRandomPhrases = ("Want","Try","Terrific","Superb","Super","Nice","Looking for","Great Deal","Give","Gift Item","Best Ever","Best","Affordable","A fabulous gift")
                        num = randrange(0,12)
                    
                        print("{} #{} #{} #{}, Sold over {}, {} sold in last hour http://rover.ebay.com/rover/1/711-53200-19255-0/1?icep_ff3=2&pub=5575036718&toolid=10001&campid={}&customid=&icep_item={}&ipn=psmain&icep_vectorid=229466&kwid=902099&mtid=824&kw=lg ".format(MyRandomPhrases[num],title,str(tag1),str(tag2),mostSold,mostSoldinhours,yourebaycampaignID,itemID))
                        
                        
                        print('NEXT PROCESS')

                        print(errorcount)
                        print(countpromotions)
                        print(totalprice)

                        if errorcount >= 3:
                            sys.exit("TOO MANY ERRORS HAVE OCCURED. EXITING TWEETER. HAVE A NICE DAY! CHECK YOUR TWITTER ISN'T LOCKED. TRY AGAIN IN A FEW HOURS.")

                        if countpromotions >= 7:

                            break
    except Exception as ex:
        print(ex)
        print("MOVING TO NEXT PROCESS")
        pass


def MyEbayTestCode():
    try:


        global extractMostPopular
        
        
        myset_data = []
        
        for fruit in extractMostPopular:

            pear = unescape2(fruit)
            print(pear)
            print('{:8} {:8} {:5} {:10} {:14} {:10}'.format('Watch', 'Sold', 'Hour', 'Price','ItemID','Title'))
            linkforprocessing = ("http://rest.ebay.com/epn/v1/find/item.rss?keyword="+ pear +"&sortOrder=BestMatch&programid=1&campaignid=5337424366&toolid=10039&minPrice="+ priceinput +".0&listingType1=All&topRatedSeller=true&feedType=rss&lgeo=1")

            f = requests.get(linkforprocessing)
             
            
            websiteData = f.text
            
            

            extractEbayItemIDFromWebsite = re.findall ('<guid>(.+?)</guid>', websiteData)



            global price
            for itemID in extractEbayItemIDFromWebsite:
            
                title = AnalyseAnItemIDForTitle(itemID)
                price = AnalyseItemIDForPrice(itemID)
                title = BMP(title)
                title = unescape6(title)
                title = title[0:81]
                price = unescape5(price)
                price = pricefix(price)
                price = float(price)
                mostSold = AnalyseAnItemID(itemID)
                mostSoldinhours = AnalyseAnItemIDForhours(itemID)
                watchers = AnalyseAnItemIDForWatchers(itemID)
                

                print('{:*^8} {:_>8} {:*^5} {:_>10.2f} {:_^14} {:_<10}'.format(watchers, mostSold, mostSoldinhours, price, itemID, title))


                if price >= float(priceinput) and watchers >= int(watchersinput) and mostSold >= int(mostSoldinput) and mostSoldinhours >= int(mostSoldinhoursvar):
                    
                    
                    global countpromotions
                    global totalprice
                    increment()
                    

                    if mostSold > 2000 and mostSoldinhours > 2:

                        if itemID in myset_data:
                            print("Skipping Duplicate")
                            break
                        myset_data.append(itemID)

                        title = str(title)
                        price = '{0:.2f}'.format(price)
                        price = str(price)
                        mostSold = str(mostSold)
                        mostSoldinhours = str(mostSoldinhours)
                        watchers = str(watchers)
                        itemID = str(itemID)

                        now = datetime.now() # current date and time
                        keyword = pear
                        dateforspreadsheet = now.strftime("%m/%d/%Y")
                        yearforspreadsheet = now.strftime("%Y")
                        monthforspreadsheet = now.strftime("%B")
                        timeforspreadsheet = now.strftime("%H:%M:%S")

                        
                        print(str(title), file=open(myfile, 'a'))
                        fileforappending.close()
                        print("appending ITEM ID: " + itemID + " : " + title + " data to file")
                        html_file_update(your_html_file_name,title,yourebaycampaignID,price,itemID)



                        myebayurl = "http://rover.ebay.com/rover/1/711-53200-19255-0/1?icep_ff3=2&pub=5575036718&toolid=10001&campid=" + yourebaycampaignID +"&customid=&icep_item=" + itemID + "&ipn=psmain&icep_vectorid=229466&kwid=902099&mtid=824&kw=lg"


                        row_increment(itemID,title,price,watchers,mostSold,mostSoldinhours,pear,dateforspreadsheet,monthforspreadsheet,timeforspreadsheet,myebayurl)

                    
                        

                        #CHECK EBAY ITEM CONVERSION URL AND MOSTSOLDINHOUR TEXT

                        
                        mystr = title
                        wordList = re.sub("[^\w]", " ",  mystr).split()

                        NEWWORDLIST = (sample(wordList, 2))

                        
                        
                        
                        tag1 = NEWWORDLIST[0]
                        tag2 = NEWWORDLIST[1]

                        

                        MyRandomPhrases = ("Thousands sold","Hot Product","Trending","Selling Fast","Fantastic offer","Why wait?","Trusted seller!","Top seller!","Want","Try","Terrific","Superb","Super","Nice","Looking for","Great Deal","Give","Gift Item","Best Ever","Best","Affordable","A fabulous gift")
                        MyRandomPrices = ("Yours for","Only","Just","Buy it now!","On sale","For","Get it now","Buy now","Going for","Best price","Bargain","Selling for")
                        num = randrange(0,20)
                        num2 = randrange(0,10)
                    
                        print("{} #{} #{} #{}, Sold over {}, {} sold in last hour. {} ${} http://rover.ebay.com/rover/1/711-53200-19255-0/1?icep_ff3=2&pub=5575036718&toolid=10001&campid={}&customid=&icep_item={}&ipn=psmain&icep_vectorid=229466&kwid=902099&mtid=824&kw=lg ".format(MyRandomPhrases[num],title,str(tag1),str(tag2),mostSold,mostSoldinhours,MyRandomPrices[num2],price,yourebaycampaignID,itemID))
                      
                        print('NEXT PROCESS')
                        print(errorcount)
                        print(countpromotions)
                        print(totalprice)

                        if errorcount >= 3:
                            sys.exit("TOO MANY ERRORS HAVE OCCURED. EXITING TWEETER. HAVE A NICE DAY! CHECK YOUR TWITTER ISN'T LOCKED. TRY AGAIN IN A FEW HOURS.")
                            
                        if countpromotions >= 7:

                            break
     


                    elif mostSold > 0 and mostSoldinhours > 0:

                        if itemID in myset_data:
                            print("Skipping Duplicate")
                            break
                        myset_data.append(itemID)

                        title = str(title)
                        price = '{0:.2f}'.format(price)
                        price = str(price)
                        mostSold = str(mostSold)
                        mostSoldinhours = str(mostSoldinhours)
                        watchers = str(watchers)
                        itemID = str(itemID)

                        now = datetime.now() # current date and time
                        keyword = pear
                        dateforspreadsheet = now.strftime("%m/%d/%Y")
                        yearforspreadsheet = now.strftime("%Y")
                        monthforspreadsheet = now.strftime("%B")
                        timeforspreadsheet = now.strftime("%H:%M:%S")

                        
                        print(str(title), file=open(myfile, 'a'))
                        fileforappending.close()
                        print("appending ITEM ID: " + itemID + " : " + title + " data to file")
                        html_file_update(your_html_file_name,title,yourebaycampaignID,price,itemID)


                        myebayurl = "http://rover.ebay.com/rover/1/711-53200-19255-0/1?icep_ff3=2&pub=5575036718&toolid=10001&campid=" + yourebaycampaignID +"&customid=&icep_item=" + itemID + "&ipn=psmain&icep_vectorid=229466&kwid=902099&mtid=824&kw=lg"


                        row_increment(itemID,title,price,watchers,mostSold,mostSoldinhours,pear,dateforspreadsheet,monthforspreadsheet,timeforspreadsheet,myebayurl)

                    

                        #CHECK EBAY ITEM CONVERSION URL AND MOSTSOLDINHOUR TEXT

                        
                        mystr = title
                        wordList = re.sub("[^\w]", " ",  mystr).split()

                        NEWWORDLIST = (sample(wordList, 2))

                        
                        
                        
                        tag1 = NEWWORDLIST[0]
                        tag2 = NEWWORDLIST[1]

                        

                        MyRandomPhrases = ("Want","Try","Terrific","Superb","Super","Nice","Looking for","Great Deal","Give","Gift Item","Best Ever","Best","Affordable","A fabulous gift")
                        MyRandomPrices = ("Yours for","Only","Just","Buy it now!","On sale","For","Get it now","Buy now","Going for","Best price","Bargain","Selling for")
                        num = randrange(0,12)
                        num2 = randrange(0,10)
                    
                        print("{} #{} #{} #{}, Sold over {}, {} sold in last hour. {} ${} http://rover.ebay.com/rover/1/711-53200-19255-0/1?icep_ff3=2&pub=5575036718&toolid=10001&campid={}&customid=&icep_item={}&ipn=psmain&icep_vectorid=229466&kwid=902099&mtid=824&kw=lg ".format(MyRandomPhrases[num],title,str(tag1),str(tag2),mostSold,mostSoldinhours,MyRandomPrices[num2],price,yourebaycampaignID,itemID))
                        print('NEXT PROCESS')
                        print(errorcount)
                        print(countpromotions)
                        print(totalprice)

                        if errorcount >= 3:
                            sys.exit("TOO MANY ERRORS HAVE OCCURED. EXITING TWEETER. HAVE A NICE DAY! CHECK YOUR TWITTER ISN'T LOCKED. TRY AGAIN IN A FEW HOURS.")
                            
                        if countpromotions >= 7:

                            break
                        



                    elif mostSold == 0:

                        if "apple" or "Apple" in title:


                            if itemID in myset_data:
                                print("Skipping Duplicate")
                                break
                            myset_data.append(itemID)

                
                            title = str(title)
                            price = '{0:.2f}'.format(price)
                            price = str(price)
                            mostSold = str(mostSold)
                            mostSoldinhours = str(mostSoldinhours)
                            watchers = str(watchers)
                            itemID = str(itemID)

                            now = datetime.now() # current date and time
                            keyword = pear
                            dateforspreadsheet = now.strftime("%m/%d/%Y")
                            yearforspreadsheet = now.strftime("%Y")
                            monthforspreadsheet = now.strftime("%B")
                            timeforspreadsheet = now.strftime("%H:%M:%S")

                            
                            print(str(title), file=open(myfile, 'a'))
                            fileforappending.close()
                            print("appending ITEM ID: " + itemID + " : " + title + " data to file")
                            html_file_update(your_html_file_name,title,yourebaycampaignID,price,itemID)

                            myebayurl = "http://rover.ebay.com/rover/1/711-53200-19255-0/1?icep_ff3=2&pub=5575036718&toolid=10001&campid=" + yourebaycampaignID +"&customid=&icep_item=" + itemID + "&ipn=psmain&icep_vectorid=229466&kwid=902099&mtid=824&kw=lg"
                            row_increment(itemID,title,price,watchers,mostSold,mostSoldinhours,pear,dateforspreadsheet,monthforspreadsheet,timeforspreadsheet,myebayurl)


                            
                            mystr = title
                            wordList = re.sub("[^\w]", " ",  mystr).split()

                            NEWWORDLIST = (sample(wordList, 2))

                            
                            
                            
                            tag1 = NEWWORDLIST[0]
                            tag2 = NEWWORDLIST[1]
                            #RANDOM TECH PHRASES
                            MyRandomPhrases = ("Want","Try","Terrific","Superb","Super","Nice","Looking for","Great Deal","Give","Gift Item","Best Ever","Best","Affordable","A fabulous gift")
                            MyRandomPrices = ("Yours for","Only","Just","Buy it now!","On sale","For","Get it now","Buy now","Going for","Best price","Bargain","Selling for")
                            num = randrange(0,12)
                            num2 = randrange(0,10)
                        
                            print("{} #{} #{} #{}, {} Watchers, {} ${} http://rover.ebay.com/rover/1/711-53200-19255-0/1?icep_ff3=2&pub=5575036718&toolid=10001&campid={}&customid=&icep_item={}&ipn=psmain&icep_vectorid=229466&kwid=902099&mtid=824&kw=lg ".format(MyRandomPhrases[num],title,str(tag1),str(tag2),watchers,MyRandomPrices[num2],price,yourebaycampaignID,itemID))
                          
                            print('NEXT PROCESS')
                            print(errorcount)
                            print(countpromotions)
                            print(totalprice)


                            
                            
                            if errorcount >= 3:
                                sys.exit("TOO MANY ERRORS HAVE OCCURED. EXITING TWEETER. HAVE A NICE DAY! CHECK YOUR TWITTER ISN'T LOCKED. TRY AGAIN IN A FEW HOURS.")

                            if countpromotions >= 7:
                                
                                break
                    

                        elif "antique" or "Antique" or "antique," or "Antique," in title:


                            if itemID in myset_data:
                                print("Skipping Duplicate")
                                break
                            myset_data.append(itemID)



                            title = str(title)
                            price = '{0:.2f}'.format(price)
                            price = str(price)
                            mostSold = str(mostSold)
                            mostSoldinhours = str(mostSoldinhours)
                            watchers = str(watchers)
                            itemID = str(itemID)

                            now = datetime.now() # current date and time
                            keyword = pear
                            dateforspreadsheet = now.strftime("%m/%d/%Y")
                            yearforspreadsheet = now.strftime("%Y")
                            monthforspreadsheet = now.strftime("%B")
                            timeforspreadsheet = now.strftime("%H:%M:%S")


                            
                            print(str(title), file=open(myfile, 'a'))
                            fileforappending.close()
                            print("appending ITEM ID: " + itemID + " : " + title + " data to file")
                            html_file_update(your_html_file_name,title,yourebaycampaignID,price,itemID)


                            myebayurl = "http://rover.ebay.com/rover/1/711-53200-19255-0/1?icep_ff3=2&pub=5575036718&toolid=10001&campid=" + yourebaycampaignID +"&customid=&icep_item=" + itemID + "&ipn=psmain&icep_vectorid=229466&kwid=902099&mtid=824&kw=lg"
                            row_increment(itemID,title,price,watchers,mostSold,mostSoldinhours,pear,dateforspreadsheet,monthforspreadsheet,timeforspreadsheet,myebayurl)
                            
                            
                            mystr = title
                            wordList = re.sub("[^\w]", " ",  mystr).split()

                            NEWWORDLIST = (sample(wordList, 2))

                            
                            
                            
                            tag1 = NEWWORDLIST[0]
                            tag2 = NEWWORDLIST[1]


                            #RANDOM ANTIQUE
                            MyRandomPhrases = ("Rare","Collectable","Fine","Great find","Collector peice","Seeking","Looking for","Great Deal","Give","Gift Item","Best Ever","Best","Affordable","A fabulous gift")
                            MyRandomPrices = ("Yours for","Only","Just","Buy it now!","On sale","For","Get it now","Buy now","Going for","Best price","Bargain","Selling for")
                            num = randrange(0,12)
                            num2 = randrange(0,10)
                        
                            print("{} #{} #{} #{}, {} Watchers, {} ${} http://rover.ebay.com/rover/1/711-53200-19255-0/1?icep_ff3=2&pub=5575036718&toolid=10001&campid={}&customid=&icep_item={}&ipn=psmain&icep_vectorid=229466&kwid=902099&mtid=824&kw=lg ".format(MyRandomPhrases[num],title,str(tag1),str(tag2),watchers,MyRandomPrices[num2],price,yourebaycampaignID,itemID))

                            
                            print('NEXT PROCESS')
                            print(errorcount)
                            print(countpromotions)
                            print(totalprice)
  
                            if errorcount >= 3:
                                sys.exit("TOO MANY ERRORS HAVE OCCURED. EXITING TWEETER. HAVE A NICE DAY! CHECK YOUR TWITTER ISN'T LOCKED. TRY AGAIN IN A FEW HOURS.")

                            if countpromotions >= 7:
                                
                                break

                        
                        else:


                            if itemID in myset_data:
                                print("Skipping Duplicate")
                                break
                            myset_data.append(itemID)



                            title = str(title)
                            price = '{0:.2f}'.format(price)
                            price = str(price)
                            mostSold = str(mostSold)
                            mostSoldinhours = str(mostSoldinhours)
                            watchers = str(watchers)
                            itemID = str(itemID)

                            now = datetime.now() # current date and time
                            keyword = pear
                            dateforspreadsheet = now.strftime("%m/%d/%Y")
                            yearforspreadsheet = now.strftime("%Y")
                            monthforspreadsheet = now.strftime("%B")
                            timeforspreadsheet = now.strftime("%H:%M:%S")



                            
                            print(str(title), file=open(myfile, 'a'))
                            fileforappending.close()
                            print("appending ITEM ID: " + itemID + " : " + title + " data to file")
                            html_file_update(your_html_file_name,title,yourebaycampaignID,price,itemID)


                            workbook = load_workbook(filename="analysis.xlsx")
                            sheet = workbook.active

                            myebayurl = "http://rover.ebay.com/rover/1/711-53200-19255-0/1?icep_ff3=2&pub=5575036718&toolid=10001&campid=" + yourebaycampaignID +"&customid=&icep_item=" + itemID + "&ipn=psmain&icep_vectorid=229466&kwid=902099&mtid=824&kw=lg"

                            row_increment(itemID,title,price,watchers,mostSold,mostSoldinhours,pear,dateforspreadsheet,monthforspreadsheet,timeforspreadsheet,myebayurl)
                        
                            workbook.save(filename="analysis.xlsx")
                            
                            mystr = title
                            wordList = re.sub("[^\w]", " ",  mystr).split()

                            NEWWORDLIST = (sample(wordList, 2))

                            
                            
                            
                            tag1 = NEWWORDLIST[0]
                            tag2 = NEWWORDLIST[1]



                            MyRandomPhrases = ("Want","Try","Terrific","Superb","Super","Nice","Looking for","Great Deal","Give","Gift Item","Best Ever","Best","Affordable","A fabulous gift")
                            MyRandomPrices = ("Yours for","Only","Just","Buy it now!","On sale","For","Get it now","Buy now","Going for","Best price","Bargain","Selling for")
                            num = randrange(0,12)
                            num2 = randrange(0,10)
                        
                            print("{} #{} #{} #{}, {} Watchers, {} ${} http://rover.ebay.com/rover/1/711-53200-19255-0/1?icep_ff3=2&pub=5575036718&toolid=10001&campid={}&customid=&icep_item={}&ipn=psmain&icep_vectorid=229466&kwid=902099&mtid=824&kw=lg ".format(MyRandomPhrases[num],title,str(tag1),str(tag2),watchers,MyRandomPrices[num2],price,yourebaycampaignID,itemID))
                           
                            print('NEXT PROCESS')
                            print(errorcount)
                            print(countpromotions)
                            print(totalprice)

                            if errorcount >= 3:
                                sys.exit("TOO MANY ERRORS HAVE OCCURED. EXITING TWEETER. HAVE A NICE DAY! CHECK YOUR TWITTER ISN'T LOCKED. TRY AGAIN IN A FEW HOURS.")

                            if countpromotions >= 7:
                                
                                break

                    elif mostSold >= 0 and mostSoldinhours  == 0:


                        if itemID in myset_data:
                            print("Skipping Duplicate")
                            break
                        myset_data.append(itemID)

                        
                        title = str(title)
                        price = '{0:.2f}'.format(price)
                        price = str(price)
                        mostSold = str(mostSold)
                        mostSoldinhours = str(mostSoldinhours)
                        watchers = str(watchers)
                        itemID = str(itemID)

                        now = datetime.now() # current date and time
                        keyword = pear
                        dateforspreadsheet = now.strftime("%m/%d/%Y")
                        yearforspreadsheet = now.strftime("%Y")
                        monthforspreadsheet = now.strftime("%B")
                        timeforspreadsheet = now.strftime("%H:%M:%S")


                        
                        print(str(title), file=open(myfile, 'a'))
                        fileforappending.close()
                        print("appending ITEM ID: " + itemID + " : " + title + " data to file")
                        html_file_update(your_html_file_name,title,yourebaycampaignID,price,itemID)

                        myebayurl = "http://rover.ebay.com/rover/1/711-53200-19255-0/1?icep_ff3=2&pub=5575036718&toolid=10001&campid=" + yourebaycampaignID +"&customid=&icep_item=" + itemID + "&ipn=psmain&icep_vectorid=229466&kwid=902099&mtid=824&kw=lg"


                        row_increment(itemID,title,price,watchers,mostSold,mostSoldinhours,pear,dateforspreadsheet,monthforspreadsheet,timeforspreadsheet,myebayurl)

                    

                        #CHECK EBAY ITEM CONVERSION URL AND MOSTSOLDINHOUR TEXT

                        
                        mystr = title
                        wordList = re.sub("[^\w]", " ",  mystr).split()

                        NEWWORDLIST = (sample(wordList, 2))

                        
                        
                        
                        tag1 = NEWWORDLIST[0]
                        tag2 = NEWWORDLIST[1]

                        MyRandomPhrases = (myhashtags + "Want","Try","Terrific","Superb","Super","Nice","Looking for","Great Deal","Give","Gift Item","Best Ever","Best","Affordable","A fabulous gift")
                        num = randrange(0,13)
                    
                        print("{} #{} #{} #{}, {} Watchers, http://rover.ebay.com/rover/1/711-53200-19255-0/1?icep_ff3=2&pub=5575036718&toolid=10001&campid={}&customid=&icep_item={}&ipn=psmain&icep_vectorid=229466&kwid=902099&mtid=824&kw=lg ".format(MyRandomPhrases[num],title,str(tag1),str(tag2),watchers,yourebaycampaignID,itemID))
                     
                        print('NEXT PROCESS')

                        print(errorcount)
                        print(countpromotions)
                        print(totalprice)
                        
                        if errorcount >= 3:
                            sys.exit("TOO MANY ERRORS HAVE OCCURED. EXITING TWEETER. HAVE A NICE DAY! CHECK YOUR TWITTER ISN'T LOCKED. TRY AGAIN IN A FEW HOURS.")
                        

                        if countpromotions >= 7:

                            break
                                            
                    
                    elif mostSold == 0 and mostSoldinhours  == 0:


                        if itemID in myset_data:
                            print("Skipping Duplicate")
                            break
                        myset_data.append(itemID)
                        
                        title = str(title)
                        price = '{0:.2f}'.format(price)
                        price = str(price)
                        mostSold = str(mostSold)
                        mostSoldinhours = str(mostSoldinhours)
                        watchers = str(watchers)
                        itemID = str(itemID)

                        now = datetime.now() # current date and time
                        keyword = pear
                        dateforspreadsheet = now.strftime("%m/%d/%Y")
                        yearforspreadsheet = now.strftime("%Y")
                        monthforspreadsheet = now.strftime("%B")
                        timeforspreadsheet = now.strftime("%H:%M:%S")


                        
                        print(str(title), file=open(myfile, 'a'))
                        fileforappending.close()
                        print("appending ITEM ID: " + itemID + " : " + title + " data to file")
                        html_file_update(your_html_file_name,title,yourebaycampaignID,price,itemID)

                        myebayurl = "http://rover.ebay.com/rover/1/711-53200-19255-0/1?icep_ff3=2&pub=5575036718&toolid=10001&campid=" + yourebaycampaignID +"&customid=&icep_item=" + itemID + "&ipn=psmain&icep_vectorid=229466&kwid=902099&mtid=824&kw=lg"


                        row_increment(itemID,title,price,watchers,mostSold,mostSoldinhours,pear,dateforspreadsheet,monthforspreadsheet,timeforspreadsheet,myebayurl)

                        #CHECK EBAY ITEM CONVERSION URL AND MOSTSOLDINHOUR TEXT

                        
                        mystr = title
                        wordList = re.sub("[^\w]", " ",  mystr).split()

                        NEWWORDLIST = (sample(wordList, 2))

                        
                        
                        
                        tag1 = NEWWORDLIST[0]
                        tag2 = NEWWORDLIST[1]

                        MyRandomPhrases = (myhashtags + "Want","Try","Terrific","Superb","Super","Nice","Looking for","Great Deal","Give","Gift Item","Best Ever","Best","Affordable","A fabulous gift")
                        num = randrange(0,13)
                    
                        print("{} #{} #{} #{}, {} Watchers, http://rover.ebay.com/rover/1/711-53200-19255-0/1?icep_ff3=2&pub=5575036718&toolid=10001&campid={}&customid=&icep_item={}&ipn=psmain&icep_vectorid=229466&kwid=902099&mtid=824&kw=lg ".format(MyRandomPhrases[num],title,str(tag1),str(tag2),watchers,yourebaycampaignID,itemID))
                        
                        print('NEXT PROCESS')

                        print(errorcount)
                        print(countpromotions)
                        print(totalprice)
                        
                        if errorcount >= 3:
                            sys.exit("TOO MANY ERRORS HAVE OCCURED. EXITING TWEETER. HAVE A NICE DAY! CHECK YOUR TWITTER ISN'T LOCKED. TRY AGAIN IN A FEW HOURS.")
                        

                        if countpromotions >= 7:

                            break
                        

        
           
                    else:

                        if itemID in myset_data:
                            print("Skipping Duplicate")
                            break
                        myset_data.append(itemID)
            
                        title = str(title)
                        price = '{0:.2f}'.format(price)
                        price = str(price)
                        mostSold = str(mostSold)
                        mostSoldinhours = str(mostSoldinhours)
                        watchers = str(watchers)
                        itemID = str(itemID)

                        now = datetime.now() # current date and time
                        keyword = pear
                        dateforspreadsheet = now.strftime("%m/%d/%Y")
                        yearforspreadsheet = now.strftime("%Y")
                        monthforspreadsheet = now.strftime("%B")
                        timeforspreadsheet = now.strftime("%H:%M:%S")

                        
                        print(str(title), file=open(myfile, 'a'))
                        fileforappending.close()
                        print("appending ITEM ID: " + itemID + " : " + title + " data to file")
                        html_file_update(your_html_file_name,title,yourebaycampaignID,price,itemID)

                        myebayurl = "http://rover.ebay.com/rover/1/711-53200-19255-0/1?icep_ff3=2&pub=5575036718&toolid=10001&campid=" + yourebaycampaignID +"&customid=&icep_item=" + itemID + "&ipn=psmain&icep_vectorid=229466&kwid=902099&mtid=824&kw=lg"

                        row_increment(itemID,title,price,watchers,mostSold,mostSoldinhours,pear,dateforspreadsheet,monthforspreadsheet,timeforspreadsheet,myebayurl)

                        #CHECK EBAY ITEM CONVERSION URL AND MOSTSOLDINHOUR TEXT

                        mystr = title
                        wordList = re.sub("[^\w]", " ",  mystr).split()

                        NEWWORDLIST = (sample(wordList, 2))

                        
                        
                        
                        tag1 = NEWWORDLIST[0]
                        tag2 = NEWWORDLIST[1]
                        



                        MyRandomPhrases = ("Want","Try","Terrific","Superb","Super","Nice","Looking for","Great Deal","Give","Gift Item","Best Ever","Best","Affordable","A fabulous gift")
                        num = randrange(0,12)
                    
                        print("{} #{} #{} #{}, Sold over {}, {} sold in last hour http://rover.ebay.com/rover/1/711-53200-19255-0/1?icep_ff3=2&pub=5575036718&toolid=10001&campid={}&customid=&icep_item={}&ipn=psmain&icep_vectorid=229466&kwid=902099&mtid=824&kw=lg ".format(MyRandomPhrases[num],title,str(tag1),str(tag2),mostSold,mostSoldinhours,yourebaycampaignID,itemID))
                        
                        
                        print('NEXT PROCESS')

                        print(errorcount)
                        print(countpromotions)
                        print(totalprice)

                        if errorcount >= 3:
                            sys.exit("TOO MANY ERRORS HAVE OCCURED. EXITING TWEETER. HAVE A NICE DAY! CHECK YOUR TWITTER ISN'T LOCKED. TRY AGAIN IN A FEW HOURS.")

                        if countpromotions >= 7:

                            break
    except Exception as ex:
        print(ex)
        print("MOVING TO NEXT PROCESS")
        pass




myfile = input("Before we begin write a file name ending with .txt , this is where itemids which meet your requirement will be saved  ...")
fileforappending = open(myfile, "w")
print(myfile + " has been created")
fileforappending.close()
print(myfile + " has been created and is ready for appending data.")
print("You will find your file here: ", os.getcwd(), "\\" , myfile)


yourebaycampaignID = input("Input your eBay CampaignID Number:")
while len(yourebaycampaignID) < 10:
    
    if yourebaycampaignID == "":
        print("eBay CampaignID == NULL, default campaignID 5338595494 will be used")
        yourebaycampaignID = "5338595494"
        break

    elif len(yourebaycampaignID) < 10:
        print("It appears you have made an error, check length of your ebay campaign ID")
        yourebaycampaignID = input("Input your eBay CampaignID Number:")

print("Your ebay campaign ID is:", yourebaycampaignID) 

 


priceinput = input("Enter your minimum price:__make this high to selectively post high ticket items to increase commission payments__don't be greedy :)")
watchersinput = input ("Enter minimum watchers__remember watch counts can be faked__but its good to have items with a lot of interest__")
mostSoldinput = input("Enter minimum most sold__sales really indicate an items popularity__enter minimum past sales__")
mostSoldinhoursvar = input("Enter minimum amount sold recently__")
myhashtags = input("Enter a single hashtag, it will be randomised with Call to actions e.g #TopOffer")



print('{:8} {:5} {:10} {:14} {:10}'.format(mostSoldinput, mostSoldinhoursvar, priceinput,'ItemID','Title'))


choose = input("Do you want to use hot trends from the internet? - YES \n Or upload your own file? - OWN")

while choose != "YES" or "OWN":
    if choose == "YES":
        MyEbayTestCode()
    elif choose == "OWN":
        ownfile()
    else:
        choose = input("Do you want to use hot trends from the internet? - YES \n Or upload your own file? - OWN")

